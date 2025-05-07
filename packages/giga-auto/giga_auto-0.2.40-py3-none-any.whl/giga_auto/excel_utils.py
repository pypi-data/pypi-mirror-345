import os
import tempfile
from contextlib import contextmanager

import pandas as pd
from openpyxl import load_workbook
from openpyxl.utils.exceptions import InvalidFileException


def modified_excel_data(**data):
    """
    修改excel文件中的单个数据
    data = {
    "file": "文件路径",
    "sheet_name": "表名",
    "item_name": "A1",
    "item_value": "value"
    }
    """
    try:
        wb = load_workbook(data['file'])
        sheet = wb[data['sheet_name']]
        sheet[data['item_name']] = data['item_value']
        wb.save(data['file'])
    except InvalidFileException:
        raise InvalidFileException("无法打开文件：文件格式不正确或已损坏。")
    except FileNotFoundError:
        raise FileNotFoundError("文件未找到，请检查文件路径是否正确。")
    except Exception as e:
        raise Exception(f"修改excel文件遇到未知错误：{e}")


def modified_excel_datas(**data):
    """
    修改excel文件中的多个数据
    data = {
    "file": "文件路径",
    "sheet_name": "表名",
    "modified_data": {
        "A1": "value1",
        "B1": "value2",
        "C1": "value3"
    }
    }
    """
    try:
        wb = load_workbook(data['file'])
        sheet = wb[data['sheet_name']]
        modified_data = data['modified_data']
        for key, value in modified_data.items():
            sheet[key] = value
        wb.save(data['file'])
    except InvalidFileException:
        raise InvalidFileException("无法打开文件：文件格式不正确或已损坏。")
    except FileNotFoundError:
        raise FileNotFoundError("文件未找到，请检查文件路径是否正确。")
    except Exception as e:
        raise Exception(f"修改excel文件遇到未知错误：{e}")


@contextmanager
def download_excel(response, filename: str = "downloaded_file.xlsx"):
    """
    上下文管理器：保存 Excel 响应内容到本地，并在退出时自动清理

    :param response: `requests` 的响应对象，需确保 response.content 是 Excel 文件
    :param filename: 自定义保存的文件名（可选，默认 "downloaded_file.xlsx"）
    :return: 返回本地 Excel 文件路径
    """
    if response.status_code != 200:
        raise RuntimeError(f"下载失败，状态码：{response.status_code}")

    # 获取当前项目的临时目录
    temp_dir = tempfile.gettempdir()
    file_path = os.path.join(temp_dir, filename)

    # 保存文件
    with open(file_path, "wb") as f:
        f.write(response.content)

    try:
        yield file_path  # 返回 Excel 文件路径
    finally:
        if os.path.exists(file_path):
            os.remove(file_path)  # 退出时自动删除文件


def read_excel_rows(file_path: str, num_rows: int = 1, sheet_name: str = ""):
    """
    读取 Excel 文件的前 num_rows 行数据
    :param file_path: Excel 文件路径
    :param num_rows: 需要读取的行数（默认 1 行）
    :sheet_name:
    :return: 前 num_rows 行的数据列表
    """
    # 如果 sheet_name 为空，则读取第一个 sheet
    df = pd.read_excel(file_path, sheet_name=sheet_name or 0, header=None)
    if df.empty:
        raise ValueError("Excel 文件为空")
    num_rows = min(num_rows, len(df))  # 防止索引超出范围
    return df.iloc[:num_rows].values.tolist()  # 返回前 num_rows 行数据


def get_sheet_names(file_path):
    """ 读取 Excel 文件中的所有 sheet 名称 """
    with pd.ExcelFile(file_path) as xls:
        return xls.sheet_names


def save_excel_datas(**data):
    """
    重新保存excel文件中的多个数据
    data = {
    "file": "文件路径",
    "sheet_name": "表名",
    "modified_data": {
        "A1": "value1",
        "B1": "value2",
        "C1": "value3"
    }
    }
    """
    import xlsxwriter
    try:
        # 创建新的工作簿
        workbook = xlsxwriter.Workbook(data['file'])
        sheet = workbook.add_worksheet(data['sheet_name'])

        # 修改数据
        modified_data = data['modified_data']
        for cell, value in modified_data.items():
            col = ord(cell[0].upper()) - ord('A')
            row = int(cell[1:]) - 1
            sheet.write(row, col, value)

        # 保存文件
        workbook.close()

    except Exception as e:
        raise Exception(f"修改excel文件遇到未知错误：{e}")