import pathlib
import re
import shutil
import os
import contextlib
import shutil
from pathlib import Path

import pandas as pd
import xlwings as xw
import openpyxl
from openpyxl import load_workbook
from contextlib import contextmanager
from .stringManager import StringBaba
class FileManagement:
    def __init__(self):
        pass

    def add_prefix(self,filename,file_type):
        pattern = r'[\u4e00-\u9fa5]+'
        matches = re.findall(pattern, filename)[0]
        return f"{matches}.{file_type}"
    def copy_files(self,src_dir, dest_dir, target_files, rename=False,file_type="xls"):
        for target_file in target_files:
            source_path = os.path.join(src_dir, target_file)
            destination_file = self.add_prefix(target_file,file_type) if rename else target_file
            destination_path = os.path.join(dest_dir, destination_file)
            if os.path.exists(source_path):
                shutil.copy(source_path, destination_path)
                print(f"File {target_file} copied from {source_path} to {destination_path}")
            else:
                print(f"Source file {target_file} not found in the latest folder.")

    def find_latest_folder(self,base_dir):
        folders = [f for f in Path(base_dir).iterdir() if f.is_dir()]
        if not folders:
            return None
        latest_folder = max(folders, key=os.path.getctime)
        return latest_folder

    def copy_file_simple(self,source_path,destination_path):
        shutil.copy(source_path, destination_path)

    def create_new_folder(self,folder_name):
        # 创建文件夹
        os.makedirs(folder_name, exist_ok=True)
        print(f"文件夹 '{folder_name}' 创建成功")

    def delete_folder_or_file(self,path):
        """
        删除指定的文件或文件夹。

        :param path: 文件或文件夹的路径
        """
        try:
            if os.path.isfile(path):
                # 如果是文件，删除文件
                os.remove(path)
                print(f"文件 '{path}' 已删除。")
            elif os.path.isdir(path):
                # 如果是文件夹，删除文件夹及其内容
                shutil.rmtree(path)
                print(f"文件夹 '{path}' 及其内容已删除。")
            else:
                print(f"路径 '{path}' 不存在。")
        except Exception as e:
            print(f"删除 '{path}' 时出错: {e}")

class ExcelHandler:
    def __init__(self, file_name):
        self.file_name = file_name
        self.wb = load_workbook(self.file_name)

    def excel_write(self, sheet_name, results, start_row, start_col, end_row, end_col):
        try:
            sheet = self.wb[sheet_name]
            for i, row in enumerate(range(start_row, end_row + 1)):
                for j, value in enumerate(range(start_col, end_col + 1)):
                    sheet.cell(row=row, column=value, value=results[i][j])
            print("Results have been written!")
            self.wb.save(self.file_name)
        except Exception as e:
            print(e)

    def excel_read(self, sheet_name, start_row, start_col, end_row, end_col):
        try:
            sheet = self.wb[sheet_name]
            values = [
                [sheet.cell(row=row, column=col).value for col in range(start_col, end_col + 1)]
                for row in range(start_row, end_row + 1)
            ]
            print("Results have been read!")
            return values
        except Exception as e:
            print(e)

    def excel_save_as(self, file_name2):
        try:
            self.wb.save(file_name2)
            print("The file has been saved as " + str(file_name2))
        except Exception as e:
            print(e)

    def excel_quit(self):
        try:
            self.wb.close()
        except Exception as e:
            print(e)

    @staticmethod
    def fast_write(sheet_name, results, start_row, start_col, end_row=0, end_col=0, re=0, xl_book=None):
        if re == 0:
            end_row = len(results) + start_row - 1
            end_col = len(results[0]) + start_col - 1
        elif re == 1:
            pass
        xl_book.excel_write(sheet_name, results, start_row=start_row, start_col=start_col, end_row=end_row, end_col=end_col)

class OpenExcel:
    def __init__(self, openfile, savefile=None):
        self.openfile = openfile
        self.savefile = savefile

    @contextmanager
    def my_open(self):
        print(f"Opening Excel file: {self.openfile}")
        wb = eExcel(file_name=self.openfile)
        yield wb
        wb.excel_save_as(self.savefile)

    @contextmanager
    def open_save_Excel(self):
        try:
            app = xw.App(visible=False)
            wb = app.books.open(self.openfile)
        except:
            app.quit()
        yield wb
        try:
            # 刷新全部数据连接和查询
            wb.api.RefreshAll()
            wb.save(self.savefile)
        finally:
            app.quit()

    def file_show(self,filter=[]):
        app = xw.App(visible=False)
        wb = app.books.open(self.openfile)
        wbsn=wb.sheet_names
        app.quit()
        if filter or filter==[""]:
            wbsn=StringBaba(wbsn).filter_string_list(filter)
        return wbsn

class ExcelOperation():
    def __init__(self, input_file,output_folder):
        self.input_file = input_file
        self.output_folder= output_folder
    def split_table(self):
        #拆分表格
        excel_file = pd.ExcelFile(self.input_file)

        # 遍历每个工作表
        for sheet_name in excel_file.sheet_names:
            # 读取当前工作表的数据
            df = pd.read_excel(self.input_file, sheet_name=sheet_name)
            # 定义输出文件名，使用工作表的名称
            output_file = f'{sheet_name}.xlsx'
            # 将当前工作表的数据保存为新的 Excel 文件
            df.to_excel(f"{self.output_folder}/{output_file}", index=False)

class eExcel():
    def __init__(self, file_name=None):
        self.file_name = file_name
        if not pathlib.Path(file_name).exists():
            self.create_new_excel(file_name)
        self.wb = openpyxl.load_workbook(file_name)
        self.ws = self.wb.active

    @staticmethod
    def create_new_excel(file_name):
        wb = openpyxl.Workbook()
        sheet = wb.active
        sheet.title = 'sheet1'  # 设置工作表的名称为sheet1
        wb.save(file_name)

    def create_new_sheet(self,ws):
        self.wb.create_sheet(ws)

    def excel_write(self,ws, results, start_row, start_col, end_row, end_col):
        ws=self.wb[ws]
        for i, row in enumerate(range(start_row, end_row + 1)):
            for j, value in enumerate(range(start_col, end_col + 1)):
                ws.cell(row=row, column=value, value=results[i][j])

    def excel_read(self, start_row, start_col, end_row, end_col):
        valueA = [
            [self.ws.cell(row=row, column=col).value for col in range(start_col, end_col + 1)]
            for row in range(start_row, end_row + 1)
        ]
        return valueA

    def excel_save_as(self, file_name2):
        self.wb.save(file_name2)

    def fast_write(self, ws, results, sr, sc, er=0, ec=0, re=0,wb=None):
        if re == 0:
            er = len(results) + sr - 1
            ec = len(results[0]) + sc - 1
        elif re == 1:
            pass
        wb.excel_write(ws, results, start_row=sr, start_col=sc, end_row=er, end_col=ec)
