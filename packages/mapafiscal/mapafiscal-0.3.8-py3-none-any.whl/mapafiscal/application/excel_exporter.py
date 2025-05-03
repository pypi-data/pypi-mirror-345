import openpyxl
from openpyxl.styles import PatternFill, Border, Side, Alignment, Font
from datetime import datetime
from copy import copy
from pathlib import Path
from mapafiscal.domain.aggregates import MapaFiscal

path = Path(__file__).parent


# Estilos padrão para escrita no excel
default_font = Font(name='Calibri',
                size=8,
                bold=False,
                italic=False,
                vertAlign=None,
                underline='none',
                strike=False,
                color='FF000000')

default_pattern_fill = PatternFill(fill_type=None,
                start_color='FFFFFFFF',
                end_color='FF000000')

default_border = Border(left=Side(style='thin'), 
                        right=Side(style='thin'), 
                        top=Side(style='thin'), 
                        bottom=Side(style='thin'))               

default_alignment=Alignment(horizontal='general',
                    vertical='bottom',
                    text_rotation=0,
                    wrap_text=False,
                    shrink_to_fit=False,
                    indent=0)

default_number_format = 'General'           


class ExcelExporter():    
           
    def __init__(self, mapa_fiscal: MapaFiscal):
        self.mapa_fiscal = mapa_fiscal   
        
    def __fill_line(self, worksheet, from_line: int, data):
        for col, value in enumerate(data):
            worksheet.cell(row=from_line, column=col+1).value = value    
            
    def __copy_line_style(self, worksheet, from_line: int, n_lines: int, n_cols: int):
        for i in range(from_line + 1, from_line + n_lines):
            for j in range(1, n_cols + 1):
                worksheet.cell(row=i, column=j)._style = copy(worksheet.cell(row=i-1, column=j)._style)              
           
    def to_excel(self, file_path: str):
       
        arquivo = openpyxl.load_workbook(f'{path}\\templates\\mapafiscal.xlsx')
        
        # ----- Aba 0 - Versão -----
        
        aba = arquivo["0 - VERSAO"]
        offset = 3
        
        ufs = list(set([cenario_fiscal.uf_destino for cenario_fiscal in self.mapa_fiscal.cenarios]))        
        classes_fiscais = list(set([classe_fiscal.ncm for classe_fiscal in self.mapa_fiscal.classes_fiscais]))
        
        values = ['v1.0', 
                  datetime.now().strftime("%d/%m/%Y"), 
                  ', '.join(classes_fiscais), 
                  ', '.join(ufs), 
                  'versão inicial']
            
        self.__fill_line(aba, offset, values)
        self.__copy_line_style(aba, offset, len(self.mapa_fiscal.classes_fiscais), len(values))
        
        # ----- Aba 1 - NCM -----
        
        aba = arquivo["1 - NCM"]
        offset = 3        
        
        for classe_fiscal in self.mapa_fiscal.classes_fiscais:
            
            values = ["v1.0",
                      classe_fiscal.ncm,
                      classe_fiscal.descricao,
                      classe_fiscal.ipi.descricao_tipi,
                      classe_fiscal.cest,
                      classe_fiscal.segmento,
                      classe_fiscal.fabricante_equiparado,                      
                      classe_fiscal.origem,
                      classe_fiscal.icms.cst.codigo,
                      classe_fiscal.icms.aliquota / 100.0,
                      classe_fiscal.icms.fcp / 100.0,
                      classe_fiscal.ipi.cst.codigo,
                      classe_fiscal.ipi.aliquota / 100.0,
                      classe_fiscal.pis.cst.codigo,
                      classe_fiscal.pis.aliquota / 100.0,
                      classe_fiscal.cofins.cst.codigo,
                      classe_fiscal.cofins.aliquota / 100.0,
                      self.mapa_fiscal.regime_tributario.descricao,
                      self.mapa_fiscal.regime_especial.descricao if self.mapa_fiscal.regime_especial else '',
                      classe_fiscal.icms.fundamento_legal]
            
            self.__fill_line(aba, offset, values)            
            self.__copy_line_style(aba, offset, len(self.mapa_fiscal.classes_fiscais), len(values))          
            offset += 1
        
        # ----- Aba 2 - ST -----    

        aba = arquivo["2 - ST"]
        offset = 3
        
        for classe_st in self.mapa_fiscal.classes_st:
            
            values = ["v1.0",
                      classe_st.ncm,
                      classe_st.cest,
                      classe_st.descricao_cest,
                      self.mapa_fiscal.uf_origem,
                      classe_st.uf_destino,
                      classe_st.mva_original / 100.0,
                      classe_st.mva_ajustada_4 / 100.0,
                      classe_st.mva_ajustada_12 / 100.0,
                      classe_st.fundamento_legal]
            
            self.__fill_line(aba, offset, values)
            self.__copy_line_style(aba, offset, len(self.mapa_fiscal.classes_st), len(values))
            offset += 1
        
        # ----- Aba 3 - Cenarios -----
        
        aba = arquivo["3 - CENARIOS"]        
        offset = 3
        
        for cenario in self.mapa_fiscal.cenarios:
            
            values = ["v1.0",
                      cenario.grupo,
                      cenario.natureza_operacao.descricao,
                      cenario.classe_fiscal.descricao,
                      cenario.classe_fiscal.ncm,
                      cenario.classe_fiscal.cest,
                      cenario.uf_origem,
                      cenario.uf_destino,
                      cenario.tipo_cliente.descricao,
                      cenario.finalidade.descricao,
                      'Sim' if cenario.incide_icms else 'Não',
                      'Sim' if cenario.incide_icms_st else 'Não',
                      'Sim' if cenario.incide_icms_st == False and cenario.incide_difal else 'Não',
                      'Sim' if cenario.incide_icms_st and cenario.incide_difal else 'Não',
                      'Sim' if cenario.regime_especial else 'Não',]
            
            self.__fill_line(aba, offset, values)
            self.__copy_line_style(aba, offset, len(self.mapa_fiscal.cenarios), len(values))
            offset += 1
        
        # ----- Aba 4 - Mapa Fiscal -----
        
        aba = arquivo["4 - MAPA FISCAL"]        
        offset = 3

        for operacao in self.mapa_fiscal.operacoes:
            values = ["v1.0",
                      operacao.cenario.grupo,
                      operacao.cenario.natureza_operacao.descricao,
                      operacao.cenario.uf_origem,
                      operacao.cenario.uf_destino,
                      operacao.cenario.classe_fiscal.descricao,
                      operacao.cenario.classe_fiscal.origem,
                      operacao.cenario.classe_fiscal.ncm,
                      operacao.cenario.classe_fiscal.cest,
                      operacao.cenario.finalidade.descricao,
                      operacao.cenario.tipo_cliente.descricao,
                      'ICMS',
                      operacao.cfop_saida,
                      operacao.cfop_entrada,
                      operacao.cenario.cst_icms.codigo,
                      operacao.reducao_bc_icms / 100.0,
                      operacao.aliq_icms_operacao / 100.0,
                      operacao.aliq_icms_uf_destino / 100.0,
                      operacao.aliq_fcp_uf_destino / 100.0,
                      operacao.difal_icms / 100.0,
                      'ICMS ST',
                      operacao.reducao_bc_icms_st / 100.0,
                      operacao.mva_st / 100.0,
                      operacao.difal_icms_st / 100.0,
                      "IPI",
                      operacao.cenario.cst_ipi.codigo,
                      operacao.aliq_ipi / 100.0,
                      "PIS",
                      operacao.cenario.cst_pis_cofins.codigo,
                      operacao.aliq_pis / 100.0,
                      "COFINS",
                      operacao.cenario.cst_pis_cofins.codigo,
                      operacao.aliq_cofins / 100.0,         
                      operacao.fundamento_legal]
            
            self.__fill_line(aba, offset, values)
            self.__copy_line_style(aba, offset, len(self.mapa_fiscal.operacoes), len(values))         
            offset += 1
            
        arquivo.save(file_path)