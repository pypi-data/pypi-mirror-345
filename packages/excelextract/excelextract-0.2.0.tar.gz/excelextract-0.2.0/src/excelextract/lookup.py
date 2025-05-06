
import re

from openpyxl.utils import column_index_from_string, get_column_letter

from .tokens import applyTokenReplacement

def resolveLookups(wb, elements = [], unprocessedDefinitions = [], currentElement = {}):
    if len(unprocessedDefinitions) == 0:
        elements.append(currentElement)
    else:
        loopDefinition = unprocessedDefinitions[0]

        if not isinstance(loopDefinition, dict):
            raise ValueError(f"Invalid loop definition: {loopDefinition}")
        
        if "operation" not in loopDefinition:
            raise ValueError(f"Missing 'operation' in loop definition: {loopDefinition}")

        operation = loopDefinition["operation"].lower()
        if operation not in ["loopsheets", "findrow", "findcolumn", "looprows", "loopcolumns"]:
            raise ValueError(f"Invalid loop operation '{operation}' in definition: {loopDefinition}")
        
        if "token" not in loopDefinition:
            raise ValueError(f"Missing 'token' in loop definition: {loopDefinition}")
        token = loopDefinition["token"]
        
        loopElements = []
        
        if operation == "loopsheets":
            if "regex" not in loopDefinition:
                raise ValueError(f"Missing 'regex' in loop definition: {loopDefinition}")
            sheetRegex = loopDefinition["regex"]
            matchingSheets = [sheet.title for sheet in wb.worksheets if re.search(sheetRegex, sheet.title)]
            loopElements = matchingSheets

        elif operation == "findrow" or operation == "findcolumn":
            if "match" not in loopDefinition:
                raise ValueError(f"Missing 'match' in loop definition: {loopDefinition}")
            match = loopDefinition["match"]

            if type(match) == str:
                match = [match]

            if "sheet" not in loopDefinition:
                raise ValueError(f"Missing 'sheet' in loop definition: {loopDefinition}")
            sheet = applyTokenReplacement(loopDefinition["sheet"], currentElement)

            if operation == "findrow":
                if "column" not in loopDefinition:
                    raise ValueError(f"Missing 'column' in loop definition: {loopDefinition}")
                searchSlice = applyTokenReplacement(loopDefinition["column"], currentElement)
            else:
                if "row" not in loopDefinition:
                    raise ValueError(f"Missing 'row' in loop definition: {loopDefinition}")
                searchSlice = applyTokenReplacement(loopDefinition["row"], currentElement)
            
            data = [str(cell.value) if cell.value is not None else "" for cell in wb[sheet][searchSlice]]
            indices = [i for i, s in enumerate(data) if s in match]

            if len(indices) == 0:
                raise ValueError(f"No matches found for '{match}' in sheet '{sheet}' and search slice '{searchSlice}'")

            if "mode" not in loopDefinition or loopDefinition["mode"] == "first":
                indices = [indices[0]]
            elif loopDefinition["mode"] == "last":
                indices = [indices[-1]]
            elif loopDefinition["mode"] == "all":
                pass
            elif type(loopDefinition["mode"]) == int:
                indices = [indices[loopDefinition["mode"]]]
            else:
                raise ValueError(f"Invalid mode '{loopDefinition['mode']}' in definition: {loopDefinition}")

            offset = loopDefinition.get("offset", 0)
            indices = [i + offset + 1 for i in indices] # +1 to convert to 1-based index

            if operation == "findrow":
                loopElements = indices
            else:
                loopElements = [get_column_letter(i) for i in indices]

        elif operation == "looprows" or operation == "loopcolumns":
            if "start" not in loopDefinition:
                raise ValueError(f"Missing 'start' in loop definition: {loopDefinition}")
            start = applyTokenReplacement(loopDefinition["start"], currentElement)
            if operation == "loopcolumns":
                start = column_index_from_string(start)
            else:
                start = int(start)

            if "end" in loopDefinition and "count" in loopDefinition:
                raise ValueError("Cannot specify both 'end' and 'count' in loop definition")

            if "end" in loopDefinition:
                end = applyTokenReplacement(loopDefinition["end"], currentElement)
                if operation == "loopcolumns":
                    end = column_index_from_string(end)
            elif "count" in loopDefinition:
                count = applyTokenReplacement(loopDefinition["count"], currentElement)
                end = start + count - 1
            elif "untilNoMatch" in loopDefinition and loopDefinition["untilNoMatch"]:
                raise ValueError("untilNoMatch is not implemented yet")
            else:
                raise ValueError("Must specify either 'end' or 'count' in loop definition")

            stride = loopDefinition.get("stride", 1)
            startOffset = loopDefinition.get("startOffset", 0)
            if startOffset != 0:
                start += startOffset
            endOffset = loopDefinition.get("endOffset", 0)
            if endOffset != 0:
                end += endOffset

            if end < start:
                raise ValueError(f"Start index {start} is greater than stop index {end} in definition: {loopDefinition}")

            indices = list(range(start, end + 1, stride))
            if operation == "looprows":
                loopElements = indices
            else:
                loopElements = [get_column_letter(i) for i in indices]

        if len(loopDefinition) == 0:
            raise ValueError("Loop definition is empty")

        for i in range(len(loopElements)):
            copy = currentElement.copy()
            copy[token] = loopElements[i]
            resolveLookups(wb, elements, unprocessedDefinitions[1:], copy)
