import logging

from decimal import Decimal
from enum import Enum
from typing import Any, Iterable, Optional

from openpyxl import load_workbook
from openpyxl.cell import Cell

from ofxstatement.plugin import Plugin
from ofxstatement.parser import StatementParser
from ofxstatement.statement import (
    BankAccount,
    Currency,
    Statement,
    StatementLine,
    generate_transaction_id,
    recalculate_balance,
)

logging.basicConfig(level=logging.DEBUG)
logger = logging.getLogger("BBVA")

TYPE_MAPPING = {
    "Pago con tarjeta": "POS",
    "Bizum": "PAYMENT",
    "Transferencia realizada": "XFER",
    "Transferencia recibida": "XFER",
}

TYPE_MAPPING_PREFIXES = {
    "Abono Bonificación pack": "FEE",
    "Abono ": "DIRECTDEP",
    "Adeudo ": "DIRECTDEBIT",
    "Retirada de efectivo": "ATM",
}


class Fields(Enum):
    VALUE_DATE = "F.Valor"
    DATE = "Fecha"
    CONCEPT = "Concepto"
    MOVEMENT = "Movimiento"
    AMOUNT = "Importe"
    CURRENCY = "Divisa"
    BALANCE = "Disponible"
    DESCRIPTION = "Observaciones"


class BBVAParser(StatementParser[str]):
    date_format = "%d/%m/%Y"

    def __init__(self, filename: str) -> None:
        super().__init__()
        self.filename = filename

        logging.debug("Loading %s", self.filename)
        self._bank_account = None
        self._start_row = 0
        self._start_column = 0
        self._ws = load_workbook(self.filename).active
        self._fields_to_row = {}

    def parse(self) -> Statement:
        found = False

        fields_values = [f.value.lower() for f in Fields]
        for row in self._ws:
            for cell in row:
                if isinstance(cell.value, str) and (
                    cell.value.lower() in fields_values
                ):
                    start_row = cell.row
                    start_column = cell.col_idx - 1
                    found = True
                    break

            if found:
                break

        if not found:
            raise ValueError(f"No '{Fields.VALUE_DATE.value}' cell found")

        logging.debug(
            "Statement table start cell found at %s",
            self._ws[start_row][start_column].coordinate,
        )

        for field in Fields:
            for cell in self._ws[start_row][start_column:]:
                if cell.value == field.value:
                    self._fields_to_row[field] = cell.col_idx - start_column - 1
                    break

        logging.debug("Statement table mapping are %s", f"{self._fields_to_row}")

        if not [Fields.DATE] & self._fields_to_row.keys():
            raise ValueError("No date column found")

        if not [Fields.AMOUNT] & self._fields_to_row.keys():
            raise ValueError("No amount column found")

        if not [Fields.CONCEPT] & self._fields_to_row.keys():
            raise ValueError("No concept column")

        self._start_row = start_row + 1
        self._start_column = start_column

        for row in self._ws.iter_rows(
            min_row=self._start_row, min_col=self._start_column
        ):
            for cell in row:
                balance = self.get_field_record(
                    row[self._start_column :], Fields.BALANCE
                )
                if balance:
                    self.statement.start_balance = self.parse_value(balance, "amount")

        if self.statement.account_id:
            logger.debug("Account ID: %s", self.statement.account_id)

        if self.statement.currency:
            logger.debug("Currency: %s", self.statement.currency)

        self._bank_account = BankAccount(
            bank_id="MICSITM1XXX", acct_id=self.statement.account_id
        )

        statement = super().parse()

        for sl in statement.lines:
            if not statement.currency or statement.currency == sl.currency:
                statement.currency = sl.currency
            else:
                statement.currency = None
                break

        recalculate_balance(statement)

        if self.statement.start_balance:
            logger.debug("Start balance: %0.2f", self.statement.start_balance)

        if self.statement.end_balance:
            logger.debug("End balance: %0.2f", self.statement.end_balance)

        return statement

    def split_records(self) -> Iterable[Iterable[Cell]]:

        cells = []
        row = self._start_row
        while True:
            line_contents = self._ws[row][self._start_column :]

            if not any(cell.value for cell in line_contents):
                break

            cells.append(line_contents)
            row += 1

        return cells

    def get_field_record(self, cells: Iterable[Cell], field: Fields) -> Any:
        if field not in self._fields_to_row.keys():
            return None

        return cells[self._fields_to_row[field]].value

    def strip_spaces(self, string: str):
        return " ".join(string.strip().split())

    def parse_value(self, value: Optional[Any], field: str) -> Any:
        if field == "amount":
            if isinstance(value, (int, float)):
                return Decimal(value)

        elif field == "date":
            if not isinstance(value, str):
                return value

        elif field == "memo":
            try:
                return self.strip_spaces(value.split(" - ", 1)[1])
            except:
                pass

        return super().parse_value(value, field)

    def get_transaction_type(self, concept, movement) -> tuple:
        transaction_type = TYPE_MAPPING.get(concept)
        if transaction_type:
            return (transaction_type, concept)

        transaction_type = TYPE_MAPPING.get(movement)
        if transaction_type:
            return (transaction_type, movement)

        for prefix, transaction_type in TYPE_MAPPING_PREFIXES.items():
            if concept.startswith(prefix):
                return (transaction_type, concept)

        logger.warning("Mapping not found for '%s' - '%s'", concept, movement)
        return ("OTHER", None)

    def parse_record(self, cells: Iterable[Cell]) -> StatementLine:
        stat_line = StatementLine(
            date=self.parse_value(self.get_field_record(cells, Fields.DATE), "date"),
            amount=self.parse_value(
                self.get_field_record(cells, Fields.AMOUNT), "amount"
            ),
        )

        concept = self.get_field_record(cells, Fields.CONCEPT)
        movement = self.get_field_record(cells, Fields.MOVEMENT)
        description = self.get_field_record(cells, Fields.DESCRIPTION)

        (stat_line.trntype, type_source) = self.get_transaction_type(concept, movement)

        memo_elements = []
        if type_source != concept:
            if stat_line.amount < 0:
                stat_line.payee = concept
            memo_elements.append(concept)

        if type_source != movement:
            memo_elements.append(movement)

        stat_line.memo = " - ".join(memo_elements + [description])

        currency = self.parse_value(
            self.get_field_record(cells, Fields.CURRENCY), "currency"
        )
        if currency:
            stat_line.currency = Currency(symbol=currency)

        stat_line.bank_account_to = self._bank_account
        stat_line.id = generate_transaction_id(stat_line)

        logging.debug(stat_line)
        stat_line.assert_valid()

        return stat_line


class BBVAPlugin(Plugin):
    """BBVA parser"""

    def get_parser(self, filename: str) -> BBVAParser:
        return BBVAParser(filename)
