import os
import fitz  # PyMuPDF
from docx import Document
import openpyxl
from .exceptions import *
from .context import PDFContext
from docx.document import Document as _Document
from docx.table import Table
from docx.text.paragraph import Paragraph
from docx.oxml.table import CT_Tbl
from docx.oxml.text.paragraph import CT_P

class FileHandler:
    # def __init__(self, file_path, start=None, end=None):
    #     self.file_path = file_path
    #     self.start = start
    #     self.end = end
    #     self.doc = None
    #     self.file_type = None  # 'pdf', 'docx', 'xlsx'
    #     self._validate_and_load()

    def __init__(self, file_path, start=None, end=None, sheet_name=None, sheet_start=None, sheet_end=None):
        self.file_path = file_path
        self.start = start
        self.end = end
        self.sheet_name = sheet_name
        self.sheet_start = sheet_start
        self.sheet_end = sheet_end
        self.doc = None
        self.file_type = None
        self._validate_and_load()

    def _validate_and_load(self):
        if not os.path.exists(self.file_path):
            raise FileNotFoundErrorCustom(f"File not found: {self.file_path}")

        file_lower = self.file_path.lower()

        if file_lower.endswith(".pdf"):
            self.file_type = "pdf"
            self.doc = fitz.open(self.file_path)
            self._validate_pdf_pages()

        elif file_lower.endswith(".docx"):
            self.file_type = "docx"
            self.doc = Document(self.file_path)

        elif file_lower.endswith(".xlsx"):
            self.file_type = "xlsx"
            self.doc = openpyxl.load_workbook(self.file_path)

        else:
            raise FileTypeNotSupportedError("Only PDF, DOCX, and XLSX files are supported.")

    def _validate_pdf_pages(self):
        total_pages = len(self.doc)

        if self.start is not None and not isinstance(self.start, int):
            raise ValueError("Start page must be an integer.")
        if self.end is not None and not isinstance(self.end, int):
            raise ValueError("End page must be an integer.")

        if self.start is not None and (self.start < 1 or self.start > total_pages):
            raise ValueError(f"Start page {self.start} is out of bounds. PDF has {total_pages} pages.")
        if self.end is not None and (self.end < 1 or self.end > total_pages):
            raise ValueError(f"End page {self.end} is out of bounds. PDF has {total_pages} pages.")
        if self.end is not None and self.start is None:
            raise ValueError("End page is provided but start page is missing.")
        if self.end is not None and self.start is not None and self.start > self.end:
            raise ValueError(f"End page ({self.end}) cannot be less than start page ({self.start}).")

    def load(self):
        if self.file_type == "pdf":
            self._load_pdf()
        elif self.file_type == "docx":
            self._load_docx()
        elif self.file_type == "xlsx":
            self._load_xlsx()
        else:
            raise RuntimeError("Unsupported file type loaded.")

    def _load_pdf(self):
        if self.doc is None:
            raise RuntimeError("PDF document is not loaded.")

        if self.start is not None and self.end is not None:
            start_idx = self.start - 1
            end_idx = self.end
        elif self.start is not None:
            start_idx = self.start - 1
            end_idx = self.start
        else:
            start_idx = 0
            end_idx = len(self.doc)

        text = "".join(self.doc[i].get_text() for i in range(start_idx, end_idx))

        if not text.strip():
            raise EmptyFileError("The selected pages in the PDF file are empty.")

        PDFContext.set_text(text)
  
    def iter_block_items(self, parent):
        """Yield paragraphs and tables in document order"""
        for child in parent.element.body.iterchildren():
            if isinstance(child, CT_P):
                yield Paragraph(child, parent)
            elif isinstance(child, CT_Tbl):
                yield Table(child, parent)

    def _load_docx(self):
        if self.doc is None:
            raise RuntimeError("DOCX document is not loaded.")

        full_text = []

        for block in self.iter_block_items(self.doc):
            if isinstance(block, Paragraph):
                text = block.text.strip()
                if text:
                    full_text.append(text)
            elif isinstance(block, Table):
                for row in block.rows:
                    row_text = [cell.text.strip() for cell in row.cells]
                    if any(row_text):
                        full_text.append('\t'.join(row_text))

        if not full_text:
            raise EmptyFileError("The DOCX file is empty or contains only whitespace.")

        combined_text = '\n'.join(full_text)
        PDFContext.set_text(combined_text)

    def _load_xlsx(self):
        if self.doc is None:
            raise RuntimeError("XLSX workbook is not loaded.")

        all_text = []

        for sheet in self.doc.worksheets:
            all_text.append(f"Sheet: {sheet.title}")
            for row in sheet.iter_rows(values_only=True):
                row_text = [str(cell) if cell is not None else "" for cell in row]
                if any(row_text):
                    all_text.append('\t'.join(row_text))

        combined_text = '\n'.join(all_text)

        if not combined_text.strip():
            raise EmptyFileError("The Excel file is empty or contains only whitespace.")

        PDFContext.set_text(combined_text)

    def _load_xlsx(self):
        if self.doc is None:
            raise RuntimeError("XLSX workbook is not loaded.")

        all_text = []

        sheets = self.doc.worksheets
        sheet_names = [s.title for s in sheets]

        # Load by sheet name
        if self.sheet_name:
            if self.sheet_name not in sheet_names:
                raise ValueError(f"Sheet name '{self.sheet_name}' not found in workbook.")
            sheets_to_read = [self.doc[self.sheet_name]]

        # Load by sheet number range
        elif self.sheet_start is not None or self.sheet_end is not None:
            total_sheets = len(sheets)

            if self.sheet_start is not None and not isinstance(self.sheet_start, int):
                raise ValueError("Sheet start must be an integer.")
            if self.sheet_end is not None and not isinstance(self.sheet_end, int):
                raise ValueError("Sheet end must be an integer.")

            if self.sheet_start is not None and (self.sheet_start < 1 or self.sheet_start > total_sheets):
                raise ValueError(f"Sheet start {self.sheet_start} is out of bounds. Workbook has {total_sheets} sheets.")
            if self.sheet_end is not None and (self.sheet_end < 1 or self.sheet_end > total_sheets):
                raise ValueError(f"Sheet end {self.sheet_end} is out of bounds. Workbook has {total_sheets} sheets.")
            # if self.sheet_end is not None and self.sheet_start is None:
            #     raise ValueError("Sheet end is provided but sheet start is missing.")
            if self.sheet_end is not None and self.sheet_start is not None and self.sheet_start > self.sheet_end:
                raise ValueError(f"Sheet end ({self.sheet_end}) cannot be less than sheet start ({self.sheet_start}).")

            start_idx = (self.sheet_start - 1) if self.sheet_start else 0
            end_idx = self.sheet_end if self.sheet_end else self.sheet_start

            sheets_to_read = sheets[start_idx:end_idx]

        # Load all sheets
        else:
            sheets_to_read = sheets

        for sheet in sheets_to_read:
            all_text.append(f"Sheet: {sheet.title}")
            for row in sheet.iter_rows(values_only=True):
                row_text = [str(cell) if cell is not None else "" for cell in row]
                if any(row_text):
                    all_text.append('\t'.join(row_text))

        combined_text = '\n'.join(all_text)

        if not combined_text.strip():
            raise EmptyFileError("The Excel file is empty or contains only whitespace.")

        PDFContext.set_text(combined_text)
